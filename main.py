import pandas as pd
import openpyxl
from seleniumwire import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import undetected_chromedriver as uc
from bs4 import BeautifulSoup
import time
import random
import requests
import os
from dotenv import load_dotenv
import re
from datetime import datetime

class Scraper:
    def __init__(self):
        load_dotenv()
        self.surname_number = os.getenv('SURNAME_NUMBER')
        self.ids_number = os.getenv('IDS_NUMBER')
        self.result_ids = []
        self.results = []
    def get_driver_uc(self):
        options = uc.ChromeOptions()

        # Simulate real user behavior
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--start-maximized")
        options.add_argument("--lang=en-US,en")
        
        # Spoof user agent
        options.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/116.0.0.0 Safari/537.36"
        )

        # Enable headful mode (headless = more likely to be detected)
        driver = uc.Chrome(options=options)

        driver.maximize_window()
        return driver
    def get_driver_wire(self):
        options = Options()
        options.add_argument("--start-maximized")
        #options.add_argument("--headless")  # 👉 Run browser in headless mode
        #options.add_argument("--disable-gpu")  # Optional: helps on Windows
        options.add_argument("--window-size=1920,1080")  # Optional: simulate full HD

        driver = webdriver.Chrome(options=options)
        return driver
    def get_chunk_surnames_filepath(self):
        chunk_file = f'./data/surnames_{self.surname_number}.xlsx'
        print(f'Use Surnames Chunk File: {chunk_file}')

        return chunk_file
    def get_chunk_ids_filepath(self):
        chunk_file = f'./ids/ids_{self.ids_number}.xlsx'
        print(f'Use IDs Chunk File: {chunk_file}')

        return chunk_file
    def get_surnames(self):
        surnames = []
        chunk_file = self.get_chunk_surnames_filepath()
        workbook = openpyxl.load_workbook(chunk_file)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            surnames.append(row[0])

        return surnames
    def get_ids(self):
        ids = []
        chunk_file = self.get_chunk_ids_filepath()
        workbook = openpyxl.load_workbook(chunk_file)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            ids.append({
                "id": row[0],
                "surname": row[1]
            })

        return ids
    def delay(self, min=1, max=2):
        time.sleep(random.uniform(min, max))
    def export_results_excel(self):
        df = pd.DataFrame(self.results)
        df.to_excel(f'results/results_{self.ids_number}.xlsx', index=False)
    def export_ids_excel(self):
        df = pd.DataFrame(self.result_ids)
        df.to_excel(f'ids/ids_{self.surname_number}.xlsx', index=False)
    def start_get_ids(self):
        self.driver = self.get_driver_wire()
        self.wait = WebDriverWait(self.driver, 10)
        self.surnames = self.get_surnames()

        for index, surname in enumerate(self.surnames):
            print("Row: ", index+1)
            print("Surname: ", surname)
            search_url = "https://portale.fnomceo.it/cerca-prof/index.php"
            self.driver.get(search_url)
            self.wait.until(EC.visibility_of_element_located((By.ID, "cognomeID")))
            surname_text = self.driver.find_element(By.ID, "cognomeID")
            surname_text.send_keys(surname)
            submit_button = self.driver.find_element(By.ID, "submitButtonID")
            # After submit_button.click()
            submit_button.click()
            time.sleep(1)  # Optional short wait before checking requests

            elenco_html = None
            start_time_try_surname = time.time()
            timeout = 10  # seconds

            while time.time() - start_time_try_surname < timeout:
                for request in reversed(self.driver.requests):
                    if request.response and "elenco.php" in request.url:
                        elenco_html = request.response.body.decode('utf-8', errors='replace')
                        break
                if elenco_html:
                    break
                time.sleep(0.5)  # Wait and retry

            if elenco_html:
                soup = BeautifulSoup(elenco_html, "html.parser")
                table = soup.find("table", {"id": "dataTableID"})
                if table:
                    rows = table.find("tbody").find_all("tr")
                    for row in rows:
                        tds = row.find_all("td")
                        if len(tds) < 1:
                            continue
                        id_cell = row.find("td")
                        if id_cell:
                            self.result_ids.append({
                                "id": id_cell.get_text(strip=True),
                                "surname": surname,
                            })

        print(f"🆔 Total IDs collected: {len(self.result_ids)}")
    def start_get_detail(self):
        self.ids = self.get_ids()
        
        for index, id in enumerate(self.ids):
            print("Row: ", index+1)
            url = "https://portale.fnomceo.it/cerca-prof/dettaglio.php"
            payload = {
                "id": id["id"]
            }

            headers = {
                "Content-Type": "application/x-www-form-urlencoded",
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/115.0.0.0 Safari/537.36"
                ),
                "Referer": "https://portale.fnomceo.it/cerca-prof/",
                "Origin": "https://portale.fnomceo.it"
            }

            response = requests.post(url, data=payload, headers=headers)
            # If request successful
            if not response.status_code == 200:
                print(f"❌ Request failed with status: {response.status_code}")
                break
                
            html = response.text
            soup = BeautifulSoup(html, "html.parser")
            title_tag = soup.find('h4', class_='modal-title')
            name = ""
            iscrizioni_1 = ""
            iscrizioni_2 = ""
            lauree_1 = ""
            lauree_2 = ""
            abilitazioni_1 = ""
            abilitazioni_2 = ""
            specializzazioni_1 = ""
            specializzazioni_2 = ""

            # Get and clean the text
            if title_tag:
                name = title_tag.get_text(strip=True)
            else:
                print("Name not found")

            for li in soup.find_all('li', class_='list-group-item'):
                badge = li.find('span', class_='badge')
                if badge and 'iscrizioni' in badge.get_text(strip=True).lower():
                    badge.extract()
                    text = li.get_text(strip=True)
                    text = re.sub(r'\s+', ' ', text)
                    if iscrizioni_1 == "":
                        iscrizioni_1 = text
                    elif iscrizioni_2 == "":
                        iscrizioni_2 = text 
                if badge and 'lauree' in badge.get_text(strip=True).lower():
                    badge.extract()
                    text = li.get_text(strip=True)
                    text = re.sub(r'\s+', ' ', text)
                    if lauree_1 == "":
                        lauree_1 = text
                    elif lauree_2 == "":
                        lauree_2 = text 
                if badge and 'abilitazioni' in badge.get_text(strip=True).lower():
                    badge.extract()
                    text = li.get_text(strip=True)
                    text = re.sub(r'\s+', ' ', text)
                    if abilitazioni_1 == "":
                        abilitazioni_1 = text
                    elif abilitazioni_2 == "":
                        abilitazioni_2 = text 
                if badge and 'specializzazioni' in badge.get_text(strip=True).lower():
                    badge.extract()
                    text = li.get_text(strip=True)
                    text = re.sub(r'\s+', ' ', text)
                    if specializzazioni_1 == "":
                        specializzazioni_1 = text
                    elif specializzazioni_2 == "":
                        specializzazioni_2 = text 
            

            self.results.append({
                "name": name,
                "iscrizioni_1": iscrizioni_1,
                "iscrizioni_2": iscrizioni_2,
                "lauree_1": lauree_1,
                "lauree_2": lauree_2,
                "abilitazioni_1": abilitazioni_1,
                "abilitazioni_2": abilitazioni_2,
                "specializzazioni_1": specializzazioni_1,
                "specializzazioni_2": specializzazioni_2,
                "surname": id["surname"],
                "id": id["id"],
            })
        
        print(f"🆔 Total Result collected: {len(self.results)}")

scraper = Scraper()
start_time = datetime.now()

try:
    #scraper.start_get_ids()
    scraper.start_get_detail()
except Exception as e:
    print(f"An error occurred: {e}")
finally:
    #scraper.export_ids_excel()
    scraper.export_results_excel()
    end_time = datetime.now()
    print("🔄 Start time:", start_time.strftime("%Y-%m-%d %H:%M:%S"))
    print("\n✅ End time:", end_time.strftime("%Y-%m-%d %H:%M:%S"))
    duration = end_time - start_time
    print(f"⏱ Duration: {duration}")
